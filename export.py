import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook

# Initialize Firebase Admin SDK
cred = credentials.Certificate(r"C:\Users\USER\Downloads\xact-db-key.json")
firebase_admin.initialize_app(cred)

db = firestore.client()

def export_event_data():
    root_collection = "2025"  # Root collection
    workbook = Workbook()  # Create a new Excel workbook
    event_data = {}  # Dictionary to store data for each event

    # Get all documents in the root collection
    colleges = db.collection(root_collection).stream()

    for college_doc in colleges:
        print(f"Processing college: {college_doc.id}")
        participants_ref = db.collection(f"{root_collection}/{college_doc.id}/Participants")
        participants = participants_ref.stream()

        for participant_doc in participants:
            participant = participant_doc.to_dict()

            # Ensure `events` exists and is an array
            if "events" not in participant or not isinstance(participant["events"], list):
                print(f"Skipping participant {participant_doc.id} due to invalid 'events' field.")
                continue

            # Process events
            for event in participant["events"]:
                if event not in event_data:
                    event_data[event] = []

                event_data[event].append({
                    "SNO": len(event_data[event]) + 1,
                    "ID": participant.get("id", "N/A"),
                    "Name": participant.get("name", "N/A"),
                    "Teams": ", ".join(participant.get("teamName", [])),
                    "College": participant.get("college", "N/A"),
                    "WNo": participant.get("wNo", "N/A"),
                })

    # Create worksheets for each event
    for event, participants in event_data.items():
        sheet = workbook.create_sheet(title=event)
        sheet.append(["SNO", "ID", "Name", "Teams", "College", "WNo"])  # Header row

        for participant in participants:
            sheet.append([
                participant["SNO"],
                participant["ID"],
                participant["Name"],
                participant["Teams"],
                participant["College"],
                participant["WNo"],
            ])

    # Remove default sheet if unused
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # Save the workbook
    output_path = "event_data.xlsx"
    workbook.save(output_path)
    print(f"Data exported successfully to {output_path}")

# Run the export function
if __name__ == "__main__":
    export_event_data()